
import urllib2
import re
from openpyxl import Workbook
from openpyxl import load_workbook

urls = ["https://www.liquorland.com.au/White%20Wine/?show=50", "https://www.liquorland.com.au/Red%20Wine/?show=50", "https://www.liquorland.com.au/Search?q=whisky&show=50", "https://www.liquorland.com.au/Sparkling/?show=50", "https://www.liquorland.com.au/Beer/?show=50"]
brands_list = []
prices = []
desc = []
flag = True
num = 1

link = urls[2]
print "test link %s" % link+"\&page="+str(num)
temp = link
while flag:
	init_brands = len(brands_list)
	print "init_brands %s"% init_brands
	try:
		content = urllib2.urlopen(link).read()
		print "html content is captured from the page"
		brands_list+= re.compile("\s+data-brand=\"([A-Za-z0-9\s&()'.]+)\"").findall(content)
		print "Brands are filtered from html"
		desc+= re.compile("data-producttitle=\"([A-Za-z0-9\s&()'.]+)\"\s+data-category=\"\w+").findall(content)
		print "Titles are filtered"
		prices+= re.compile("ata-price=\"(\d+\.\d+)\"\>").findall(content)
		print "cost is recorded for each bottle"
		print "In this iteration total items captured brands and prices%s,%s,%s"%(len(brands_list),len(desc), len(prices))
		print "\n ###################################################################"
		num=num+1
		link = temp+"&page="+str(num)
		print "next webpage to browse %s"% link
	except:
		print "script failed while capturing the info from page html"
		flag = False
	if len(brands_list) - init_brands > 0:
		print "new brands captured in the page"
	else:
		break
count = 1	
try:
	wb = load_workbook('liquor_sample.xlsx')
except:
	wb = Workbook()
	wb.create_sheet("liquor_sample.xlsx")
	
ws = wb.active
try:
	for i in range(1,600):
		print "%s,%s,%s,%s"%(count, brands_list[i],desc[i],prices[i]) 
		print "spreadsheet active"
		ws['I'+str(i)] = "LIQUORLAND"
		ws['J'+str(i)] = brands_list[i-1]
		print "First column populated"
		ws['K'+str(i)] = desc[i-1]
		ws['L'+str(i)] = prices[i-1]
		count =count+1
except:
	print "Total items captured"

wb.save("liquor_sample.xlsx")
