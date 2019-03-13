from selenium.webdriver.common.keys import Keys
import selenium.webdriver as webdriver
from selenium.webdriver.common.by import By
from openpyxl import Workbook
import time
import re
import random
driver = webdriver.Chrome("/home/madhureddy/myapp/chromedriver")
#driver = webdriver.Firefox()
time.sleep(5)
print "window opened===> next==> maximise"
driver.maximize_window()
time.sleep(float(str(random.randrange(2,6,1))+"."+str(random.randrange(500,1000,26))))
driver.get("https://bws.com.au/spirits/whisky")
scroll = 2000
last_height = driver.execute_script("return document.body.scrollHeight")
while True:
	driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
	time.sleep(3)
	new_height = driver.execute_script("return document.body.scrollHeight")
	if new_height == last_height:
		break
	else:
		last_height = new_height
flag = 1
detail = {}
brands_list = []
desc = []
prices = []

content = driver.find_element_by_tag_name("body").get_attribute("innerHTML")
for num in range(1,300):
	try:
		brand_xpath = "//*[@id=\"center-panel\"]/div[1]/div[3]/bws-category/div/wow-card-list/div/div[2]/div[1]/div["+str(num)+"]/wow-card/span/bws-product/div/div[2]/a/h2"
		price_xpath = "//*[@id=\"center-panel\"]/div[1]/div[3]/bws-category/div/wow-card-list/div/div[2]/div[1]/div["+str(num) +"]/wow-card/span/bws-product/div/div[3]/span[2]"
		desc_xpath = "//*[@id=\"center-panel\"]/div[1]/div[3]/bws-category/div/wow-card-list/div/div[2]/div[1]/div["+str(num)+"]/wow-card/span/bws-product/div/div[2]/div"
		value = driver.find_element_by_xpath(brand_xpath).text
		print "brand => %s" % value
		brands_list.append(value)
		value = driver.find_element_by_xpath(desc_xpath).text
		print "description => %s" % value
		desc.append(value)
		value = driver.find_element_by_xpath(price_xpath).text
		print "price => %s" % value
		prices.append(value)
		#brands = re.compile("BrandName\"\>([a-zA-Z'\s+]+)\<\/h2").findall(content)
		##desc = re.compile("Title\"\>([a-zA-Z0-9'\s+]+)\<\/div").findall(content)
		#prices = re.compile("Dollars\"\>(\d+)<\/span").findall(html)
	except:
		try:
			img_xpath = "//*[@id=\"center-panel\"]/div[1]/div[3]/bws-category/div/wow-card-list/div/div[2]/div[1]/div["+str(num)+"]/wow-card/span/wow-inspiration-card/div/div/div[1]/a/img"
			test = driver.find_element_by_xpath(img_xpath)
		except:
			break
		
print "Brands %s" % brands_list
print "Description %s" % desc
print "Prices %s" % prices
print "Brands %s " % len(brands_list)
print "prices %s " % len(prices)

for i in range(0,83):
	print "brand: %s,  Description:%s, Price:%s"% (brands_list[i], desc[i], prices[i])

try:
	wb = load_workbook("liquor_sample.xlsx")
except:
	wb = Workbook()
	wb.create_sheet("liquor_sample.xlsx")

ws = wb.active	
count = 1
try:
	for i in range(1,600):
		print "%s,%s,%s,%s"%(count, brands_list[i],desc[i],prices[i]) 
		print "spreadsheet active"
		ws['A'+str(i)] = "BWS"
		ws['B'+str(i)] = brands_list[i-1]
		print "First column populated"
		ws['C'+str(i)] = desc[i-1]
		ws['D'+str(i)] = prices[i-1]
		count =count+1
except:
	print "Total items captured"

wb.save("liquor_sample.xlsx")	
#driver.close()
