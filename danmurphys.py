from selenium.webdriver.common.keys import Keys
import selenium.webdriver as webdriver
from selenium.webdriver.common.by import By
from openpyxl import Workbook
from openpyxl import load_workbook
import time
import re
import random
driver = webdriver.Chrome("/home/madhureddy/myapp/chromedriver")
#path="/home/madhureddy/myapp/geckodriver"
#driver = webdriver.Firefox(executable_path=path)
time.sleep(5)
print "window opened===> next==> maximise"
driver.maximize_window()
time.sleep(float(str(random.randrange(2,6,1))+"."+str(random.randrange(500,1000,26))))
driver.get("https://www.danmurphys.com.au/whisky/all")
time.sleep(2)
driver.execute_script("location.reload()");

try:
	driver.find_element_by_xpath('//*[@id="set-store"]/div/div[2]/div').click()
except:
	try:
		print "crossed one try for closing the pop-up"
		driver.find_element_by_xpath('//*[@id=\"set-store\"]/div[2]/a').click()
	except:
		print "none"
		
last_height = driver.execute_script("return document.body.scrollHeight")	
driver.execute_script("window.scrollTo(0, 250);")
new_height = driver.execute_script("return document.body.scrollHeight")
if new_height-last_height != 250:
	driver.execute_script("location.reload()");
	driver.execute_script("window.scrollTo(0, 300);")
new_height = driver.execute_script("return document.body.scrollHeight")
if new_height-last_height != 300:
	print "Scroll bar didnt move down"

time.sleep(3)	
driver.find_element_by_xpath('//*[@id=\"results\"]/div[1]/div[2]/search-results/div/div/div[2]/div[2]/div[3]').click()
	
last_height = driver.execute_script("return document.body.scrollHeight")
while True:
	driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
	time.sleep(3)
	new_height = driver.execute_script("return document.body.scrollHeight")
	if new_height == last_height:
		break
	else:
		last_height = new_height

value = ''	
brands_list = []	
prices = []
desc = []
for num in range(1,125):
	if num <=4:
		brand_xpath = "//*[@id=\"results\"]/div[1]/div[2]/search-results/div/ul/div/li["+str(num)+"]/shop-product-card/div/section/div[2]/h2/a/span[1]"
		price_xpath = "//*[@id=\"results\"]/div[1]/div[2]/search-results/div/ul/div/li["+str(num)+"]/shop-product-card/div/section/div[2]/div/div[1]/product-card-view/div/div[1]/span[1]"
		desc_xpath = "//*[@id=\"results\"]/div[1]/div[2]/search-results/div/ul/div/li["+str(num)+"]/shop-product-card/div/section/div[2]/h2/a/span[2]"
	else:	
		num = num-4
		brand_xpath = "//*[@id=\"results\"]/div[1]/div[2]/search-results/div/ul/li["+str(num)+"]/shop-product-card/div/section/div[2]/h2/a/span[1]"
		price_xpath = "//*[@id=\"results\"]/div[1]/div[2]/search-results/div/ul/li["+str(num)+"]/shop-product-card/div/section/div[2]/div/div[1]/product-card-view/div/div[1]/span[1]"
		desc_xpath = "//*[@id=\"results\"]/div[1]/div[2]/search-results/div/ul/li["+str(num)+"]/shop-product-card/div/section/div[2]/h2/a/span[2]"
	try:
		#print "brand_xpath and num %s,%s"%(brand_xpath, num)
		value = driver.find_element_by_xpath(brand_xpath).text
		print "brand => %s" % value
		brands_list.append(value)
		#price_xpath = "//*[@id=\"results\"]/div[1]/div[2]/search-results/div/ul/li["+str(num)+"]/shop-product-card/div/section/div[2]/div/div[1]/product-card-view/div/div[1]/span[1]"
		#print "price_xpath and num %s,%s"%(price_xpath, num)
		value = driver.find_element_by_xpath(desc_xpath).text
		print "description => %s" % value
		desc.append(value)
	except: 
		temp = 1

	try:
		value = driver.find_element_by_xpath(price_xpath).text
		print "price value %s" % value
		if value == "MEMBER OFFER":
			price_xpath = "//*[@id=\"results\"]/div[1]/div[2]/search-results/div/ul/li["+str(num)+"]/shop-product-card/div/section/div[2]/div/div[1]/product-card-view/div/div[2]/span"
			value = driver.find_element_by_xpath(price_xpath).text
			print "Price => %s" % value
		if len(brands_list) != len(prices):
			prices.append(value)
	except:
		value = "NULL"
		prices.append(value)	
	print "total brands and prices ===> %s, %s"%(len(brands_list), len(prices))
	print "\n##################################"
print "Brands %s"% brands_list
print "Total Brands %s"% len(brands_list)
print "Titles %s"% desc
print "Prices %s" % prices
print "Total Prices %s" % len(prices)

try:
	wb = load_workbook("liquor_sample.xlsx")
	#print "Existing xlsx is opened"
except:
	wb = Workbook()
	wb.create_sheet("liquor_sample.xlsx")
	#print "created new xlsx is opened"

ws = wb.active

count = 0
try:
	for i in range(1,600):
		print "%s,%s,%s,%s"%(count, brands_list[i],desc[i],prices[i]) 
		print "spreadsheet active"
		ws['E'+str(i)] = "DAN MURPHYS"
		ws['F'+str(i)] = brands_list[i-1]
		print "First column populated"
		ws['G'+str(i)] = desc[i-1]
		ws['H'+str(i)] = prices[i-1]
		count =count+1
except:
	print "Total items captured"
#wb = load_workbook("liquor_sample.xlsx")
wb.save("liquor_sample.xlsx")
